qtd_totalEpi = 378
epi = [
    {"produto": "CAMISA OPERAÇÃO", "tamanho": "P",  "quantidade": 52},
    {"produto": "CAMISA OPERAÇÃO", "tamanho": "M",  "quantidade": 156},
    {"produto": "CAMISA OPERAÇÃO", "tamanho": "G",  "quantidade": 77},
    {"produto": "CAMISA OPERAÇÃO", "tamanho": "GG", "quantidade": 57},
    {"produto": "CAMISA POLO VERDE", "tamanho": "M", "quantidade": 5},
    {"produto": "CAMISA SOCIAL", "tamanho": "M", "quantidade": 15},
    {"produto": "CAMISA SOCIAL", "tamanho": "G", "quantidade": 16},
]

print("ESTOQUE ATUALMENTE\n")
print(f"{'TIPO':<20} {'TAMANHO':<10} {'QUANTIDADE':<10}")

for i, item in enumerate(epi):
    if i == 0:
        print(f"\nTOTAL EM ESTOQUE: {qtd_totalEpi}\n")
    print(f"{item['produto']:<20} {item['tamanho']:<10} {item['quantidade']:<10}")




retiradada_camisasP = int(input("\nQuantos colaboradores precisam da camisa P? \n"))
retiradada_camisasM = int(input("Quantos colaboradores precisam da camisa M? \n"))
retiradada_camisasG = int(input("Quantos colaboradores precisam da camisa G? \n"))
retiradada_camisasGG = int(input("Quantos colaboradores precisam da camisa GG? \n"))

retirada_luvasP = int(input("Quantos colaboradores precisam da luva(P)\n"))
retirada_luvasM = int(input("Quantos colaboradores precisam da luva(M)\n"))
retirada_luvasG = int(input("Quantos colaboradores precisam da luva(G)\n"))
retirada_luvasGG = int(input("Quantos colaboradores precisam da luva(GG)\n"))

retirada_camisaPoloP = int(input("Quantos colaboradores precisam da Polo (P)\n"))
retirada_camisaPoloM = int(input("Quantos colaboradores precisam da Polo (M)\n"))
retirada_camisaPoloG = int(input("Quantos colaboradores precisam da Polo (G)\n"))



camisa_op_ret_P  = retiradada_camisasP  * 3
camisa_op_ret_M  = retiradada_camisasM  * 3
camisa_op_ret_G  = retiradada_camisasG  * 3
camisa_op_ret_GG = retiradada_camisasGG * 3

camisa_polo_P  = retirada_camisaPoloP  * 3
camisa_polo_M  = retirada_camisaPoloM  * 3
camisa_polo_G  = retirada_camisaPoloG  * 3




retiradas = {
    "P": camisa_op_ret_P,
    "M": camisa_op_ret_M,
    "G": camisa_op_ret_G,
    "GG": camisa_op_ret_GG
}

for item in epi:
    tamanho = item["tamanho"]

    # Subtrair SOMENTE se for camisa da operação
    if item["produto"] == "CAMISA OPERAÇÃO" and tamanho in retiradas:
        item["quantidade"] -= retiradas[tamanho]

        if item["quantidade"] < 0:
            item["quantidade"] = 0

# --- SUBTRAÇÃO DAS CAMISAS POLO VERDE ---
retirada_polo = {
    "P": camisa_polo_P,
    "M": camisa_polo_M,
    "G": camisa_polo_G,
   

}

for item in epi:
    if item["produto"] == "CAMISA POLO VERDE" and item["tamanho"] in retirada_polo:
        item["quantidade"] -= retirada_polo[item["tamanho"]]

        if item["quantidade"] < 0:
            item["quantidade"] = 0



total_retirado = camisa_op_ret_P + camisa_op_ret_M + camisa_op_ret_G + camisa_op_ret_GG
quantidade_restante = qtd_totalEpi - total_retirado



print("\n===== RESULTADO DAS RETIRADAS =====\n")
print(f"Retiradas P : {camisa_op_ret_P}")
print(f"Retiradas M : {camisa_op_ret_M}")
print(f"Retiradas G : {camisa_op_ret_G}")
print(f"Retiradas GG: {camisa_op_ret_GG}")

print(f"\nTotal retirado: {total_retirado}")
print(f"Quantidade restante no estoque: {quantidade_restante}")
from datetime import datetime

data_atual = datetime.now().strftime("%d/%m/%Y %H:%M:%S")

conteudo = "ESTOQUE ATUALMENTE\n"
conteudo += f"Atualizado em: {data_atual}\n\n"
conteudo += f"{'TIPO':<20} {'TAMANHO':<10} {'QUANTIDADE':<10}\n"

for item in epi:
    conteudo += f"{item['produto']:<20} {item['tamanho']:<10} {item['quantidade']:<10}\n"


for item in epi:
    print(f"{item['produto']:<20} {item['tamanho']:<10} {item['quantidade']:<10}")

with open("estoque_epi2.txt", "w", encoding="utf-8") as arquivo:
    arquivo.write(conteudo)

print("Arquivo 'estoque_epi2.txt' gerado com sucesso!")

import os
from openpyxl import Workbook, load_workbook

caminho_excel = "estoque_epi.xlsx"


if os.path.exists(caminho_excel):
    wb = load_workbook(caminho_excel)
    if "Estoque EPI" in wb.sheetnames:
        ws = wb["Estoque EPI"]
    else:
        ws = wb.active
        ws.title = "Estoque EPI"
    print("Arquivo Excel carregado com sucesso!")
else:
    wb = Workbook()
    ws = wb.active
    ws.title = "Estoque EPI"
    ws.append(["TIPO", "TAMANHO", "QUANTIDADE"])
    print("Arquivo Excel criado, pois não existia.")


for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
    for cell in row:
        cell.value = None

ws.delete_rows(2, ws.max_row)  


for item in epi:
    ws.append([item["produto"], item["tamanho"], item["quantidade"]])


wb.save(caminho_excel)

print("Arquivo 'estoque_epi.xlsx' atualizado com sucesso!\n")


print("Arquivo 'estoque_epi.xlsx' gerado com sucesso!")


print("")

print("\n=== LISTA ATUALIZADA PARA COPIAR E COLAR NO CÓDIGO ===\n")

for item in epi:
    print(f'{{"produto": "{item["produto"]}", "tamanho": "{item["tamanho"]}", "quantidade": {item["quantidade"]}}},')
